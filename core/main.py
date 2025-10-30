import json
import logging
from datetime import datetime
from pathlib import Path
from typing import (
    Any,
    Dict,
    List,
)

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook


logging.basicConfig(
    level=logging.INFO,
    format='%(message)s',
)
logger = logging.getLogger(__name__)


def load_workbook_safe(file_path: Path) -> Workbook | None:
    try:
        return load_workbook(file_path)
    except FileNotFoundError:
        logger.error('Файл не найден')
    except InvalidFileException:
        logger.error(f"Файл {file_path} не является xlsx файлом")


def convert_date_to_string(date: datetime) -> str:
    return date.date().isoformat()


def get_data_from_excel(workbook: Workbook) -> List[Dict[str, Any]]:
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    data = []
    for row in range(2, sheet.max_row + 1):
        data_row = {}
        for col_index, cell in enumerate(headers):
            cell_value = sheet.cell(row=row, column=col_index + 1).value
            if isinstance(cell_value, datetime):
                cell_value = convert_date_to_string(cell_value)
            if cell_value is None:
                cell_value = ""
            data_row[cell] = cell_value
        data.append(data_row)
    return data


def write_to_json_file(data: List[Dict[str, Any]], json_file_path: Path) -> None:
    with open(json_file_path, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)


def main():
    excel_file_path = Path("data/tempalte_card.xlsx")
    json_file_path = Path("data/result.json")

    wb = load_workbook_safe(excel_file_path)
    if not wb:
        return

    write_to_json_file(
            data=get_data_from_excel(wb),
            json_file_path=json_file_path,
    )
    logger.info("Файл успешно конвертирован в JSON")


if __name__ == "__main__":
    main()
